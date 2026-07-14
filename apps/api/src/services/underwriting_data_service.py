"""Private-company source ingestion, canonical facts, and QoE bridge services."""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import unicodedata
import zipfile
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.db.base import now_utc
from src.models.target import Target
from src.models.underwriting_data import (
    AccountMapping,
    AnalysisRun,
    ArtifactVersion,
    CanonicalFinancialFact,
    FinancialImportException,
    FinancialReconciliation,
    QoEAdjustment,
    SourceSnapshot,
)
from src.schemas.underwriting_data import (
    AccountMappingCreate,
    AnalysisRunCreate,
    ArtifactVersionCreate,
    FinancialImportCreate,
    NormalizedFinancialRow,
    PrivateTargetCreate,
    QoEAdjustmentCreate,
    QoEAdjustmentDecision,
    SourceSnapshotCreate,
)
from src.services.common import NotFound, get_workspace_or_404, touch_status


class UnderwritingDataError(ValueError):
    """Base class for user-correctable underwriting data errors."""


class UnderwritingDataConflict(UnderwritingDataError):
    """The request conflicts with an existing version or review decision."""


MAX_XLSX_BYTES = 20 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_ARCHIVE_MEMBERS = 2_000
MAX_XLSX_DATA_ROWS = 100_000
MAX_XLSX_COLUMNS = 25
XLSX_REQUIRED_HEADERS = {
    "raw_account",
    "statement",
    "period_end",
    "period_type",
    "value",
}
XLSX_OPTIONAL_HEADERS = {
    "canonical_account",
    "period_start",
    "scale",
    "unit",
    "currency",
}


def _json_default(value: Any) -> str:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    raise TypeError(f"{type(value).__name__} is not JSON serializable")


def _json_safe(value: Any) -> Any:
    return json.loads(json.dumps(value, sort_keys=True, default=_json_default))


def content_hash(value: Any) -> str:
    """Return a stable SHA-256 digest for JSON-like inputs."""

    canonical = json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        default=_json_default,
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def normalize_account(value: str) -> str:
    """Normalize an account label without losing its human-readable source value."""

    normalized = unicodedata.normalize("NFKC", value).casefold().strip()
    normalized = normalized.replace("&", " and ")
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return " ".join(normalized.split())


def _target_for_workspace(session: Session, workspace_id: str) -> Target:
    target = session.scalar(select(Target).where(Target.workspace_id == workspace_id))
    if target is None:
        raise UnderwritingDataError(
            "Set a private target before registering or importing underwriting data"
        )
    return target


def create_private_target(
    session: Session, workspace_id: str, data: PrivateTargetCreate
) -> Target:
    """Create a real private target without requiring ticker or CIK identifiers."""

    workspace = get_workspace_or_404(session, workspace_id)
    existing = session.scalar(select(Target).where(Target.workspace_id == workspace_id))
    if existing is not None:
        raise UnderwritingDataConflict("This workspace already has a target")

    target = Target(
        workspace_id=workspace_id,
        name=data.name,
        target_type="private_company",
        ticker=None,
        cik=None,
        sector=data.sector,
        description=data.description,
        fiscal_year_end=data.fiscal_year_end,
        data_source="User-provided private company data",
        is_synthetic=False,
        financials=None,
    )
    session.add(target)
    session.flush()
    workspace.target_id = target.id
    touch_status(workspace, "in_progress")
    session.commit()
    session.refresh(target)
    return target


def _next_stream_version(
    session: Session,
    model,
    workspace_id: str,
    type_column,
    type_value: str,
):
    latest = session.scalar(
        select(model)
        .where(model.workspace_id == workspace_id, type_column == type_value)
        .order_by(model.version.desc())
        .limit(1)
    )
    return (latest.version + 1 if latest else 1), latest


def _next_source_version(
    session: Session, workspace_id: str, source_type: str, source_name: str
) -> tuple[int, SourceSnapshot | None]:
    latest = session.scalar(
        select(SourceSnapshot)
        .where(
            SourceSnapshot.workspace_id == workspace_id,
            SourceSnapshot.source_type == source_type,
            SourceSnapshot.source_name == source_name,
        )
        .order_by(SourceSnapshot.version.desc())
        .limit(1)
    )
    return (latest.version + 1 if latest else 1), latest


def register_source_snapshot(
    session: Session,
    workspace_id: str,
    data: SourceSnapshotCreate,
    *,
    actor_id: str | None = None,
) -> SourceSnapshot:
    """Register an unverified user-supplied source reference without laundering provenance."""

    get_workspace_or_404(session, workspace_id)
    target = _target_for_workspace(session, workspace_id)
    source_type = "user_registered_reference"
    source_name = data.source_name.strip()
    version, latest = _next_source_version(
        session, workspace_id, source_type, source_name
    )
    snapshot = SourceSnapshot(
        workspace_id=workspace_id,
        target_id=target.id,
        source_kind="user_input",
        source_type=source_type,
        source_name=source_name,
        version=version,
        supersedes_id=latest.id if latest else None,
        filename=data.filename,
        content_type=data.content_type,
        storage_uri=data.storage_uri,
        input_hash=data.input_hash or data.content_hash,
        content_hash=data.content_hash,
        byte_size=data.byte_size,
        record_count=data.record_count,
        status="partial",
        source_metadata=_json_safe(
            {
                "provenance_origin": "user_submitted_reference",
                "verification_status": "unverified",
                "hash_attestation": "client_asserted",
                "declared": {
                    "source_kind": data.source_kind,
                    "source_type": data.source_type,
                    "status": data.status,
                    "input_hash": data.input_hash,
                    "content_hash": data.content_hash,
                },
                "declared_metadata": data.source_metadata,
            }
        ),
        created_by=actor_id or data.created_by,
    )
    session.add(snapshot)
    session.commit()
    session.refresh(snapshot)
    return snapshot


def list_source_snapshots(session: Session, workspace_id: str) -> list[SourceSnapshot]:
    get_workspace_or_404(session, workspace_id)
    return list(
        session.scalars(
            select(SourceSnapshot)
            .where(SourceSnapshot.workspace_id == workspace_id)
            .order_by(SourceSnapshot.created_at.desc(), SourceSnapshot.version.desc())
        )
    )


def create_account_mapping(
    session: Session, workspace_id: str, data: AccountMappingCreate
) -> AccountMapping:
    get_workspace_or_404(session, workspace_id)
    normalized = normalize_account(data.raw_account)
    if not normalized:
        raise UnderwritingDataError("raw_account must contain letters or numbers")

    source_type = "management_financials"
    latest = session.scalar(
        select(AccountMapping)
        .where(
            AccountMapping.workspace_id == workspace_id,
            AccountMapping.source_type == source_type,
            AccountMapping.raw_account_normalized == normalized,
        )
        .order_by(AccountMapping.version.desc())
        .limit(1)
    )
    approved_by = data.approved_by or (data.created_by if data.status == "approved" else None)
    mapping = AccountMapping(
        workspace_id=workspace_id,
        source_type=source_type,
        raw_account=data.raw_account.strip(),
        raw_account_normalized=normalized,
        canonical_account=data.canonical_account,
        statement=data.statement,
        sign_multiplier=data.sign_multiplier,
        status=data.status,
        version=latest.version + 1 if latest else 1,
        supersedes_id=latest.id if latest else None,
        created_by=data.created_by,
        approved_by=approved_by,
        approved_at=now_utc() if approved_by else None,
    )
    session.add(mapping)
    session.commit()
    session.refresh(mapping)
    return mapping


def list_account_mappings(session: Session, workspace_id: str) -> list[AccountMapping]:
    get_workspace_or_404(session, workspace_id)
    return list(
        session.scalars(
            select(AccountMapping)
            .where(AccountMapping.workspace_id == workspace_id)
            .order_by(
                AccountMapping.raw_account_normalized,
                AccountMapping.version.desc(),
            )
        )
    )


def _approved_mapping_index(
    session: Session, workspace_id: str, source_type: str
) -> dict[str, AccountMapping]:
    mappings = session.scalars(
        select(AccountMapping)
        .where(
            AccountMapping.workspace_id == workspace_id,
            AccountMapping.source_type == source_type,
            AccountMapping.status == "approved",
        )
        .order_by(AccountMapping.version.desc())
    )
    current: dict[str, AccountMapping] = {}
    for mapping in mappings:
        current.setdefault(mapping.raw_account_normalized, mapping)
    return current


def _prepare_financial_rows(
    rows: list[NormalizedFinancialRow], mappings: dict[str, AccountMapping]
) -> tuple[list[dict], list[dict]]:
    prepared: list[dict] = []
    exceptions: list[dict] = []
    semantic_keys: set[tuple] = set()

    for index, row in enumerate(rows, start=1):
        raw_normalized = normalize_account(row.raw_account)
        if not raw_normalized:
            raise UnderwritingDataError(f"Row {index}: raw_account must contain letters or numbers")
        mapping = mappings.get(raw_normalized)
        canonical = row.canonical_account or (mapping.canonical_account if mapping else None)
        sign = mapping.sign_multiplier if mapping and not row.canonical_account else Decimal("1")
        state = "mapped_explicit" if row.canonical_account else ("mapped" if mapping else "unmapped")
        if mapping and mapping.statement != row.statement and not row.canonical_account:
            state = "mapping_exception"
            exceptions.append(
                {
                    "fact_index": index - 1,
                    "code": "mapping_statement_mismatch",
                    "severity": "high",
                    "message": (
                        f"Account '{row.raw_account}' is mapped to {mapping.statement} "
                        f"but the row declares {row.statement}"
                    ),
                    "details": {
                        "mapping_id": mapping.id,
                        "mapping_statement": mapping.statement,
                        "row_statement": row.statement,
                    },
                }
            )
        if canonical is None:
            exceptions.append(
                {
                    "fact_index": index - 1,
                    "code": "unmapped_account",
                    "severity": "medium",
                    "message": f"Account '{row.raw_account}' has no approved mapping",
                    "details": {"raw_account_normalized": raw_normalized},
                }
            )

        source_locator = row.source_locator or (
            f"{row.source_sheet}!row:{row.source_row or index}"
            if row.source_sheet
            else f"row:{row.source_row or index}"
        )
        key = (
            raw_normalized,
            row.statement,
            row.period_start,
            row.period_end,
            row.period_type,
            row.unit,
            row.currency,
        )
        if key in semantic_keys:
            raise UnderwritingDataError(
                f"Row {index}: duplicate account/period/unit; provide a dimensional account label"
            )
        semantic_keys.add(key)

        normalized_value = row.value * row.scale * sign
        provenance = {
            "provenance_origin": "user_submitted_financial_import",
            "declared_metadata": dict(row.provenance or {}),
            "import_row": index,
            "mapping_method": "explicit" if row.canonical_account else "workspace_mapping",
            "mapping_version": mapping.version if mapping and not row.canonical_account else None,
        }
        fact_data = {
            "account_mapping_id": mapping.id if mapping and not row.canonical_account else None,
            "statement": row.statement,
            "raw_account": row.raw_account,
            "raw_account_normalized": raw_normalized,
            "canonical_account": canonical,
            "mapping_state": state,
            "period_start": row.period_start,
            "period_end": row.period_end,
            "period_type": row.period_type,
            "raw_value": row.value,
            "scale_factor": row.scale,
            "value": normalized_value,
            "unit": row.unit,
            "currency": row.currency,
            "source_sheet": row.source_sheet,
            "source_row": row.source_row or index,
            "source_locator": source_locator,
            "provenance": _json_safe(provenance),
        }
        fact_data["row_hash"] = content_hash(fact_data)
        prepared.append(fact_data)
    return prepared, exceptions


def _reconcile_balance_sheets(
    prepared: list[dict], tolerance_bps: Decimal
) -> tuple[list[dict], list[dict]]:
    grouped: dict[date, list[dict]] = defaultdict(list)
    for row in prepared:
        if row["statement"] == "balance_sheet" and row["mapping_state"].startswith("mapped"):
            grouped[row["period_end"]].append(row)

    results: list[dict] = []
    exceptions: list[dict] = []
    for period_end, rows in sorted(grouped.items()):
        monetary_currencies = {
            row["currency"] for row in rows if row["unit"] == "currency" and row["currency"]
        }
        by_account: dict[str, Decimal] = defaultdict(Decimal)
        for row in rows:
            if row["canonical_account"]:
                by_account[row["canonical_account"]] += row["value"]

        assets = by_account.get("total_assets")
        liabilities_equity = by_account.get("total_liabilities_and_equity")
        if liabilities_equity is None:
            liabilities = by_account.get("total_liabilities")
            equity = by_account.get("total_equity")
            if liabilities is not None and equity is not None:
                liabilities_equity = liabilities + equity

        if len(monetary_currencies) > 1:
            status = "failed"
            difference = None
            tolerance = None
            message = f"Balance sheet for {period_end} contains multiple currencies"
            exceptions.append(
                {
                    "code": "mixed_currency_reconciliation",
                    "severity": "high",
                    "message": message,
                    "details": {"period_end": period_end.isoformat(), "currencies": sorted(monetary_currencies)},
                }
            )
        elif assets is None or liabilities_equity is None:
            status = "incomplete"
            difference = None
            tolerance = None
            missing = []
            if assets is None:
                missing.append("total_assets")
            if liabilities_equity is None:
                missing.append("total_liabilities_and_equity or total_liabilities + total_equity")
            exceptions.append(
                {
                    "code": "reconciliation_incomplete",
                    "severity": "medium",
                    "message": f"Balance sheet for {period_end} cannot be reconciled",
                    "details": {"period_end": period_end.isoformat(), "missing": missing},
                }
            )
        else:
            difference = assets - liabilities_equity
            tolerance = max(abs(assets) * tolerance_bps / Decimal("10000"), Decimal("0.01"))
            status = "passed" if abs(difference) <= tolerance else "failed"
            if status == "failed":
                exceptions.append(
                    {
                        "code": "balance_sheet_imbalance",
                        "severity": "high",
                        "message": f"Balance sheet for {period_end} does not balance",
                        "details": {
                            "period_end": period_end.isoformat(),
                            "difference": format(difference, "f"),
                            "tolerance": format(tolerance, "f"),
                        },
                    }
                )
        results.append(
            {
                "period_end": period_end,
                "assets": assets,
                "liabilities_and_equity": liabilities_equity,
                "difference": difference,
                "tolerance": tolerance,
                "status": status,
                "details": {
                    "currencies": sorted(monetary_currencies),
                    "method": "assets = liabilities + equity",
                    "tolerance_bps": format(tolerance_bps, "f"),
                },
            }
        )
    return results, exceptions


def import_financial_rows(
    session: Session,
    workspace_id: str,
    data: FinancialImportCreate,
    *,
    raw_input_hash: str | None = None,
    byte_size: int | None = None,
    actor_id: str | None = None,
) -> dict:
    """Seal a normalized import, its facts, and every validation exception atomically."""

    workspace = get_workspace_or_404(session, workspace_id)
    target = _target_for_workspace(session, workspace_id)
    source_type = "management_financials"
    source_name = data.source_name.strip()
    mappings = _approved_mapping_index(session, workspace_id, source_type)
    prepared, row_exceptions = _prepare_financial_rows(data.rows, mappings)
    reconciliations, recon_exceptions = _reconcile_balance_sheets(
        prepared, data.reconciliation_tolerance_bps
    )
    exception_specs = row_exceptions + recon_exceptions

    version, latest = _next_source_version(
        session, workspace_id, source_type, source_name
    )
    normalized_content = [
        {
            key: value
            for key, value in row.items()
            if key not in {"row_hash", "provenance", "account_mapping_id"}
        }
        for row in prepared
    ]
    input_digest = raw_input_hash or content_hash(data.model_dump(mode="json"))
    normalized_digest = content_hash(normalized_content)
    mapped_count = sum(row["mapping_state"].startswith("mapped") for row in prepared)
    status = "partial" if exception_specs else "ready"
    metadata = {
        "provenance_origin": "user_submitted_financial_import",
        "declared_source_type": data.source_type,
        "declared_metadata": dict(data.source_metadata or {}),
        "format": "normalized_financial_rows",
        "mapped_count": mapped_count,
        "unmapped_count": len(prepared) - mapped_count,
        "reconciliation_tolerance_bps": format(data.reconciliation_tolerance_bps, "f"),
    }
    snapshot = SourceSnapshot(
        workspace_id=workspace_id,
        target_id=target.id,
        source_kind="financials",
        source_type=source_type,
        source_name=source_name,
        version=version,
        supersedes_id=latest.id if latest else None,
        filename=data.filename,
        content_type=data.content_type,
        input_hash=input_digest,
        content_hash=normalized_digest,
        byte_size=byte_size,
        record_count=len(prepared),
        status=status,
        source_metadata=_json_safe(metadata),
        created_by=actor_id or data.created_by,
    )
    session.add(snapshot)
    session.flush()

    facts: list[CanonicalFinancialFact] = []
    for row in prepared:
        fact = CanonicalFinancialFact(
            workspace_id=workspace_id,
            target_id=target.id,
            source_snapshot_id=snapshot.id,
            **row,
        )
        session.add(fact)
        facts.append(fact)
    session.flush()

    reconciliation_models: list[FinancialReconciliation] = []
    for result in reconciliations:
        model = FinancialReconciliation(
            workspace_id=workspace_id,
            source_snapshot_id=snapshot.id,
            **result,
        )
        session.add(model)
        reconciliation_models.append(model)

    exception_models: list[FinancialImportException] = []
    for spec in exception_specs:
        fact_index = spec.pop("fact_index", None)
        model = FinancialImportException(
            workspace_id=workspace_id,
            source_snapshot_id=snapshot.id,
            fact_id=facts[fact_index].id if fact_index is not None else None,
            code=spec["code"],
            severity=spec["severity"],
            state="open",
            message=spec["message"],
            details=_json_safe(spec.get("details")),
        )
        session.add(model)
        exception_models.append(model)

    touch_status(workspace, "in_progress")
    session.commit()
    session.refresh(snapshot)
    for model in reconciliation_models:
        session.refresh(model)
    return {
        "snapshot": snapshot,
        "row_count": len(prepared),
        "mapped_count": mapped_count,
        "unmapped_count": len(prepared) - mapped_count,
        "open_exception_count": len(exception_models),
        "reconciliations": reconciliation_models,
    }


def preview_financial_rows(
    session: Session,
    workspace_id: str,
    data: FinancialImportCreate,
    *,
    raw_input_hash: str | None = None,
) -> dict:
    """Validate and reconcile an import without creating or changing any database record."""

    get_workspace_or_404(session, workspace_id)
    _target_for_workspace(session, workspace_id)
    source_type = "management_financials"
    source_name = data.source_name.strip()
    mappings = _approved_mapping_index(session, workspace_id, source_type)
    prepared, row_exceptions = _prepare_financial_rows(data.rows, mappings)
    reconciliations, reconciliation_exceptions = _reconcile_balance_sheets(
        prepared, data.reconciliation_tolerance_bps
    )
    exceptions = row_exceptions + reconciliation_exceptions
    version, latest = _next_source_version(session, workspace_id, source_type, source_name)
    normalized_content = [
        {
            key: value
            for key, value in row.items()
            if key not in {"row_hash", "provenance", "account_mapping_id"}
        }
        for row in prepared
    ]
    mapped_count = sum(row["mapping_state"].startswith("mapped") for row in prepared)
    return {
        "will_write": False,
        "proposed_source_version": version,
        "supersedes_source_id": latest.id if latest else None,
        "input_hash": raw_input_hash or content_hash(data.model_dump(mode="json")),
        "normalized_content_hash": content_hash(normalized_content),
        "row_count": len(prepared),
        "mapped_count": mapped_count,
        "unmapped_count": len(prepared) - mapped_count,
        "projected_status": "partial" if exceptions else "ready",
        "open_exception_count": len(exceptions),
        "exceptions": [
            {
                "code": item["code"],
                "severity": item["severity"],
                "message": item["message"],
                "details": _json_safe(item.get("details")),
                "row_number": (
                    item.get("fact_index", 0) + 1 if item.get("fact_index") is not None else None
                ),
            }
            for item in exceptions
        ],
        "reconciliations": reconciliations,
    }


def _parse_decimal(raw: str, line_number: int) -> Decimal:
    value = raw.strip()
    negative = value.startswith("(") and value.endswith(")")
    if negative:
        value = value[1:-1]
    percent = value.endswith("%")
    value = value.rstrip("%").replace(",", "").replace("$", "").strip()
    try:
        parsed = Decimal(value)
    except InvalidOperation as exc:
        raise UnderwritingDataError(f"CSV row {line_number}: invalid value '{raw}'") from exc
    if negative:
        parsed = -parsed
    if percent:
        parsed /= Decimal("100")
    return parsed


def parse_financial_csv(content: bytes, filename: str = "financials.csv") -> list[NormalizedFinancialRow]:
    """Parse a strict, header-driven CSV into the normalized import contract."""

    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise UnderwritingDataError("CSV must be UTF-8 encoded") from exc
    reader = csv.DictReader(io.StringIO(decoded))
    if not reader.fieldnames:
        raise UnderwritingDataError("CSV requires a header row")
    headers = {re.sub(r"\s+", "_", name.strip().lower()): name for name in reader.fieldnames}

    def get(source: dict, *names: str) -> str | None:
        for name in names:
            original = headers.get(name)
            if original is not None and source.get(original, "").strip():
                return source[original].strip()
        return None

    rows: list[NormalizedFinancialRow] = []
    for line_number, source in enumerate(reader, start=2):
        if not any((value or "").strip() for value in source.values()):
            continue
        value_raw = get(source, "value", "amount")
        if value_raw is None:
            raise UnderwritingDataError(f"CSV row {line_number}: value is required")
        scale_raw = (get(source, "scale", "scale_factor") or "1").casefold()
        scale_aliases = {
            "ones": Decimal("1"),
            "thousands": Decimal("1000"),
            "millions": Decimal("1000000"),
        }
        if scale_raw in scale_aliases:
            scale = scale_aliases[scale_raw]
        else:
            try:
                scale = Decimal(scale_raw)
            except InvalidOperation as exc:
                raise UnderwritingDataError(
                    f"CSV row {line_number}: invalid scale '{scale_raw}'"
                ) from exc
        statement = (get(source, "statement") or "").casefold().replace(" ", "_")
        period_type = (get(source, "period_type") or "").casefold()
        period_aliases = {"monthly": "month", "quarterly": "quarter", "annual": "year"}
        payload = {
            "raw_account": get(source, "raw_account", "account", "account_name"),
            "canonical_account": get(source, "canonical_account"),
            "statement": statement,
            "period_start": get(source, "period_start"),
            "period_end": get(source, "period_end", "date"),
            "period_type": period_aliases.get(period_type, period_type),
            "value": _parse_decimal(value_raw, line_number),
            "scale": scale,
            "unit": get(source, "unit") or "currency",
            "currency": get(source, "currency") or "USD",
            "source_sheet": get(source, "source_sheet", "sheet"),
            "source_row": int(get(source, "source_row") or line_number),
            "source_locator": get(source, "source_locator") or f"{filename}:row:{line_number}",
            "provenance": {"csv_line": line_number, "filename": filename},
        }
        try:
            rows.append(NormalizedFinancialRow.model_validate(payload))
        except (ValidationError, ValueError) as exc:
            raise UnderwritingDataError(f"CSV row {line_number}: {exc}") from exc
    if not rows:
        raise UnderwritingDataError("CSV contains no financial rows")
    return rows


def _inspect_xlsx_archive(content: bytes) -> None:
    """Reject oversized, encrypted, traversing, or active-content XLSX archives."""

    if not content or len(content) > MAX_XLSX_BYTES:
        raise UnderwritingDataError("XLSX must be non-empty and no larger than 20 MiB")
    if not content.startswith(b"PK"):
        raise UnderwritingDataError("XLSX is not a valid ZIP-based workbook")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_MEMBERS:
                raise UnderwritingDataError("XLSX contains too many archive members")
            names: set[str] = set()
            uncompressed_size = 0
            for member in members:
                name = member.filename.replace("\\", "/")
                path_parts = name.split("/")
                if name.startswith("/") or ".." in path_parts:
                    raise UnderwritingDataError("XLSX contains an unsafe archive path")
                if member.flag_bits & 0x1:
                    raise UnderwritingDataError("Encrypted XLSX workbooks are not supported")
                names.add(name)
                uncompressed_size += member.file_size
                if uncompressed_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise UnderwritingDataError(
                        "XLSX expands beyond the 100 MiB safety limit"
                    )
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise UnderwritingDataError("XLSX is missing required workbook records")
            casefolded_names = {name.casefold() for name in names}
            forbidden_parts = (
                "xl/vbaproject.bin",
                "xl/activex/",
                "xl/embeddings/",
                "xl/externallinks/",
                "xl/connections.xml",
            )
            if any(
                name == forbidden_parts[0]
                or any(name.startswith(prefix) for prefix in forbidden_parts[1:4])
                or name == forbidden_parts[4]
                for name in casefolded_names
            ):
                raise UnderwritingDataError(
                    "XLSX contains macros, embedded objects, or external connections"
                )
    except zipfile.BadZipFile as exc:
        raise UnderwritingDataError("XLSX is not a valid workbook archive") from exc


def _xlsx_header(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return re.sub(r"\s+", "_", value.strip().casefold())


def _xlsx_decimal(value: object, row_number: int, field: str) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise UnderwritingDataError(f"XLSX row {row_number}: {field} must be numeric")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        raw = value.strip()
        negative = raw.startswith("(") and raw.endswith(")")
        if negative:
            raw = raw[1:-1]
        percent = raw.endswith("%")
        raw = raw.rstrip("%").replace(",", "").replace("$", "").strip()
        try:
            parsed = Decimal(raw)
        except InvalidOperation as exc:
            raise UnderwritingDataError(
                f"XLSX row {row_number}: {field} has invalid numeric value '{value}'"
            ) from exc
        if negative:
            parsed = -parsed
        if percent:
            parsed /= Decimal("100")
        return parsed
    raise UnderwritingDataError(f"XLSX row {row_number}: {field} must be numeric")


def _xlsx_date(value: object) -> object:
    if isinstance(value, datetime):
        return value.date()
    return value


def _xlsx_cell_locator(sheet_title: str, coordinate: str) -> str:
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_.]*", sheet_title):
        sheet_ref = sheet_title
    else:
        sheet_ref = "'" + sheet_title.replace("'", "''") + "'"
    return f"{sheet_ref}!{coordinate}"


def parse_financial_xlsx(
    content: bytes, filename: str = "financials.xlsx"
) -> list[NormalizedFinancialRow]:
    """Parse the normalized first-sheet XLSX template without evaluating formulas.

    Row 1 must contain the exact required headers ``raw_account``, ``statement``,
    ``period_end``, ``period_type``, and ``value``. Optional headers are
    ``canonical_account``, ``period_start``, ``scale``, ``unit``, and ``currency``.
    Only the first worksheet is read, and it must be visible. Values must be constants;
    formula and error cells are rejected so persisted facts always represent inspected
    input.
    """

    _inspect_xlsx_archive(content)
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (InvalidFileException, KeyError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise UnderwritingDataError("XLSX could not be opened as an Excel workbook") from exc

    try:
        if not workbook.worksheets:
            raise UnderwritingDataError("XLSX contains no worksheets")
        worksheet = workbook.worksheets[0]
        if worksheet.sheet_state != "visible":
            raise UnderwritingDataError("The first XLSX worksheet must be visible")
        if worksheet.max_column > MAX_XLSX_COLUMNS:
            raise UnderwritingDataError("XLSX template exceeds the 25-column limit")
        if worksheet.max_row > MAX_XLSX_DATA_ROWS + 1:
            raise UnderwritingDataError("XLSX template exceeds the 100,000-row limit")

        header_cells = next(
            worksheet.iter_rows(min_row=1, max_row=1, max_col=worksheet.max_column),
            (),
        )
        headers: dict[str, int] = {}
        blank_header_columns: set[int] = set()
        allowed_headers = XLSX_REQUIRED_HEADERS | XLSX_OPTIONAL_HEADERS
        for column_index, cell in enumerate(header_cells, start=1):
            header = _xlsx_header(cell.value)
            if not header:
                blank_header_columns.add(column_index)
                continue
            if header not in allowed_headers:
                raise UnderwritingDataError(f"XLSX has unsupported header '{header}'")
            if header in headers:
                raise UnderwritingDataError(f"XLSX has duplicate header '{header}'")
            headers[header] = column_index
        missing = XLSX_REQUIRED_HEADERS - set(headers)
        if missing:
            raise UnderwritingDataError(
                "XLSX is missing required headers: " + ", ".join(sorted(missing))
            )

        parsed_rows: list[NormalizedFinancialRow] = []
        for cells in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            max_col=worksheet.max_column,
        ):
            if not any(cell.value is not None for cell in cells):
                continue
            row_number = cells[0].row
            if any(
                column <= len(cells) and cells[column - 1].value is not None
                for column in blank_header_columns
            ):
                raise UnderwritingDataError(
                    f"XLSX row {row_number} contains data beneath a blank header"
                )
            populated_cells = {
                header: cells[column - 1]
                for header, column in headers.items()
                if cells[column - 1].value is not None
            }
            for header, cell in populated_cells.items():
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    raise UnderwritingDataError(
                        f"XLSX row {row_number}: formulas are not allowed ({header})"
                    )
                if cell.data_type == "e":
                    raise UnderwritingDataError(
                        f"XLSX row {row_number}: error cells are not allowed ({header})"
                    )

            def cell_value(header: str, default=None):
                column = headers.get(header)
                return cells[column - 1].value if column else default

            raw_scale = cell_value("scale")
            if raw_scale is None:
                raw_scale = 1
            if isinstance(raw_scale, str) and raw_scale.strip().casefold() in {
                "ones",
                "thousands",
                "millions",
            }:
                scale = {
                    "ones": Decimal("1"),
                    "thousands": Decimal("1000"),
                    "millions": Decimal("1000000"),
                }[raw_scale.strip().casefold()]
            else:
                scale = _xlsx_decimal(raw_scale, row_number, "scale")

            statement = str(cell_value("statement") or "").strip().casefold().replace(" ", "_")
            period_type = str(cell_value("period_type") or "").strip().casefold()
            period_type = {
                "monthly": "month",
                "quarterly": "quarter",
                "annual": "year",
            }.get(period_type, period_type)
            value_cell = cells[headers["value"] - 1]
            provenance_cells = {
                header: _xlsx_cell_locator(worksheet.title, cell.coordinate)
                for header, cell in populated_cells.items()
            }
            payload = {
                "raw_account": cell_value("raw_account"),
                "canonical_account": cell_value("canonical_account"),
                "statement": statement,
                "period_start": _xlsx_date(cell_value("period_start")),
                "period_end": _xlsx_date(cell_value("period_end")),
                "period_type": period_type,
                "value": _xlsx_decimal(value_cell.value, row_number, "value"),
                "scale": scale,
                "unit": cell_value("unit", "currency") or "currency",
                "currency": cell_value("currency", "USD") or "USD",
                "source_sheet": worksheet.title,
                "source_row": row_number,
                "source_locator": _xlsx_cell_locator(worksheet.title, value_cell.coordinate),
                "provenance": {
                    "filename": filename,
                    "worksheet": worksheet.title,
                    "xlsx_row": row_number,
                    "cells": provenance_cells,
                    "template_version": "normalized-financials-v1",
                },
            }
            try:
                parsed_rows.append(NormalizedFinancialRow.model_validate(payload))
            except (ValidationError, ValueError) as exc:
                raise UnderwritingDataError(f"XLSX row {row_number}: {exc}") from exc
        if not parsed_rows:
            raise UnderwritingDataError("XLSX contains no financial rows")
        return parsed_rows
    finally:
        workbook.close()


def list_financial_facts(
    session: Session,
    workspace_id: str,
    *,
    source_snapshot_id: str | None = None,
    canonical_account: str | None = None,
    period_end: date | None = None,
    limit: int = 500,
    offset: int = 0,
) -> list[CanonicalFinancialFact]:
    get_workspace_or_404(session, workspace_id)
    query = select(CanonicalFinancialFact).where(
        CanonicalFinancialFact.workspace_id == workspace_id
    )
    if source_snapshot_id:
        query = query.where(CanonicalFinancialFact.source_snapshot_id == source_snapshot_id)
    if canonical_account:
        query = query.where(CanonicalFinancialFact.canonical_account == canonical_account)
    if period_end:
        query = query.where(CanonicalFinancialFact.period_end == period_end)
    query = query.order_by(
        CanonicalFinancialFact.period_end.desc(),
        CanonicalFinancialFact.statement,
        CanonicalFinancialFact.canonical_account,
    ).offset(offset).limit(limit)
    return list(session.scalars(query))


def list_import_exceptions(
    session: Session, workspace_id: str, source_snapshot_id: str | None = None
) -> list[FinancialImportException]:
    get_workspace_or_404(session, workspace_id)
    query = select(FinancialImportException).where(
        FinancialImportException.workspace_id == workspace_id
    )
    if source_snapshot_id:
        query = query.where(FinancialImportException.source_snapshot_id == source_snapshot_id)
    return list(session.scalars(query.order_by(FinancialImportException.created_at.desc())))


def resolve_import_exception(
    session: Session,
    workspace_id: str,
    exception_id: str,
    *,
    resolved_by: str,
) -> FinancialImportException:
    exception = session.get(FinancialImportException, exception_id)
    if exception is None or exception.workspace_id != workspace_id:
        raise NotFound(f"Financial import exception '{exception_id}' not found")
    if exception.state != "open":
        raise UnderwritingDataConflict("Financial import exception is already resolved")
    exception.state = "resolved"
    exception.resolved_by = resolved_by
    exception.resolved_at = now_utc()
    session.commit()
    session.refresh(exception)
    return exception


def list_reconciliations(
    session: Session, workspace_id: str, source_snapshot_id: str | None = None
) -> list[FinancialReconciliation]:
    get_workspace_or_404(session, workspace_id)
    query = select(FinancialReconciliation).where(
        FinancialReconciliation.workspace_id == workspace_id
    )
    if source_snapshot_id:
        query = query.where(FinancialReconciliation.source_snapshot_id == source_snapshot_id)
    return list(session.scalars(query.order_by(FinancialReconciliation.period_end.desc())))


def create_qoe_adjustment(
    session: Session, workspace_id: str, data: QoEAdjustmentCreate
) -> QoEAdjustment:
    get_workspace_or_404(session, workspace_id)
    target = _target_for_workspace(session, workspace_id)
    if data.source_snapshot_id:
        snapshot = session.get(SourceSnapshot, data.source_snapshot_id)
        if snapshot is None or snapshot.workspace_id != workspace_id:
            raise UnderwritingDataError("source_snapshot_id does not belong to this workspace")

    dedupe_payload = {
        "period_end": data.period_end,
        "bridge_layer": data.bridge_layer,
        "title": normalize_account(data.title),
        "amount": data.amount,
        "currency": data.currency,
        "source_snapshot_id": data.source_snapshot_id,
        "source_locator": data.source_locator,
    }
    dedupe_key = content_hash(dedupe_payload)
    existing = session.scalar(
        select(QoEAdjustment).where(
            QoEAdjustment.workspace_id == workspace_id,
            QoEAdjustment.dedupe_key == dedupe_key,
        )
    )
    if existing:
        raise UnderwritingDataConflict(
            f"Duplicate QoE adjustment; existing adjustment is '{existing.id}'"
        )

    adjustment = QoEAdjustment(
        workspace_id=workspace_id,
        target_id=target.id,
        source_snapshot_id=data.source_snapshot_id,
        period_start=data.period_start,
        period_end=data.period_end,
        bridge_layer=data.bridge_layer,
        title=data.title.strip(),
        description=data.description.strip(),
        category=data.category.strip(),
        amount=data.amount,
        currency=data.currency,
        is_recurring=data.is_recurring,
        is_run_rate=data.is_run_rate,
        is_cash=data.is_cash,
        owner=data.owner.strip(),
        evidence_ref=data.evidence_ref,
        source_locator=data.source_locator,
        status="proposed",
        created_by=data.created_by,
        dedupe_key=dedupe_key,
    )
    session.add(adjustment)
    session.commit()
    session.refresh(adjustment)
    return adjustment


def list_qoe_adjustments(
    session: Session, workspace_id: str, period_end: date | None = None
) -> list[QoEAdjustment]:
    get_workspace_or_404(session, workspace_id)
    query = select(QoEAdjustment).where(QoEAdjustment.workspace_id == workspace_id)
    if period_end:
        query = query.where(QoEAdjustment.period_end == period_end)
    return list(
        session.scalars(
            query.order_by(QoEAdjustment.period_end.desc(), QoEAdjustment.created_at)
        )
    )


def decide_qoe_adjustment(
    session: Session,
    workspace_id: str,
    adjustment_id: str,
    data: QoEAdjustmentDecision,
) -> QoEAdjustment:
    adjustment = session.get(QoEAdjustment, adjustment_id)
    if adjustment is None or adjustment.workspace_id != workspace_id:
        raise NotFound(f"QoE adjustment '{adjustment_id}' not found")
    if adjustment.status != "proposed":
        raise UnderwritingDataConflict("QoE adjustment already has a final decision")
    if adjustment.created_by == data.decided_by:
        raise UnderwritingDataConflict(
            "The QoE adjustment proposer cannot approve or reject the same adjustment"
        )
    if (
        data.decision == "approve"
        and adjustment.evidence_ref is None
        and (adjustment.source_snapshot_id is None or adjustment.source_locator is None)
    ):
        raise UnderwritingDataError(
            "Approved QoE adjustments require an evidence_ref or a source snapshot locator"
        )
    adjustment.status = "approved" if data.decision == "approve" else "rejected"
    adjustment.decided_by = data.decided_by
    adjustment.decided_at = now_utc()
    adjustment.decision_note = data.note.strip()
    session.commit()
    session.refresh(adjustment)
    return adjustment


_EBITDA_COMPONENTS = (
    "net_income",
    "interest_expense",
    "income_tax_expense",
    "depreciation_and_amortization",
)


def _find_reported_ebitda(
    session: Session,
    workspace_id: str,
    period_end: date | None,
    source_snapshot_id: str | None,
) -> tuple[CanonicalFinancialFact | None, Decimal | None, dict | None, list[str]]:
    query = (
        select(CanonicalFinancialFact)
        .join(SourceSnapshot, SourceSnapshot.id == CanonicalFinancialFact.source_snapshot_id)
        .where(
            CanonicalFinancialFact.workspace_id == workspace_id,
            CanonicalFinancialFact.canonical_account == "ebitda",
            CanonicalFinancialFact.mapping_state.like("mapped%"),
            CanonicalFinancialFact.unit == "currency",
            CanonicalFinancialFact.currency.is_not(None),
        )
    )
    if period_end:
        query = query.where(CanonicalFinancialFact.period_end == period_end)
    if source_snapshot_id:
        query = query.where(CanonicalFinancialFact.source_snapshot_id == source_snapshot_id)
    base = session.scalar(
        query.order_by(
            CanonicalFinancialFact.period_end.desc(),
            SourceSnapshot.created_at.desc(),
        ).limit(1)
    )
    if base:
        return base, base.value, {"method": "reported_ebitda", "fact_ids": [base.id]}, []

    component_query = (
        select(CanonicalFinancialFact)
        .join(SourceSnapshot, SourceSnapshot.id == CanonicalFinancialFact.source_snapshot_id)
        .where(
            CanonicalFinancialFact.workspace_id == workspace_id,
            CanonicalFinancialFact.canonical_account.in_(_EBITDA_COMPONENTS),
            CanonicalFinancialFact.mapping_state.like("mapped%"),
            CanonicalFinancialFact.unit == "currency",
            CanonicalFinancialFact.currency.is_not(None),
        )
    )
    if period_end:
        component_query = component_query.where(CanonicalFinancialFact.period_end == period_end)
    if source_snapshot_id:
        component_query = component_query.where(
            CanonicalFinancialFact.source_snapshot_id == source_snapshot_id
        )
    candidate = session.scalar(
        component_query.order_by(
            CanonicalFinancialFact.period_end.desc(), SourceSnapshot.created_at.desc()
        ).limit(1)
    )
    if candidate is None:
        return None, None, None, ["No reported EBITDA or complete derivation was found"]

    components = list(
        session.scalars(
            select(CanonicalFinancialFact).where(
                CanonicalFinancialFact.workspace_id == workspace_id,
                CanonicalFinancialFact.source_snapshot_id == candidate.source_snapshot_id,
                CanonicalFinancialFact.period_end == candidate.period_end,
                CanonicalFinancialFact.canonical_account.in_(_EBITDA_COMPONENTS),
                CanonicalFinancialFact.mapping_state.like("mapped%"),
                CanonicalFinancialFact.unit == "currency",
                CanonicalFinancialFact.currency.is_not(None),
            )
        )
    )
    by_account: dict[str, list[CanonicalFinancialFact]] = defaultdict(list)
    for component in components:
        by_account[component.canonical_account].append(component)
    missing = [account for account in _EBITDA_COMPONENTS if account not in by_account]
    if missing:
        return (
            candidate,
            None,
            None,
            ["Cannot derive EBITDA; missing canonical facts: " + ", ".join(missing)],
        )
    currencies = {component.currency for component in components if component.currency}
    if len(currencies) > 1:
        return candidate, None, None, ["Cannot derive EBITDA from mixed currencies"]
    value = sum(
        (
            component.value
            for account in _EBITDA_COMPONENTS
            for component in by_account[account]
        ),
        Decimal("0"),
    )
    derivation = {
        "method": "net_income + interest + income_tax + depreciation_and_amortization",
        "fact_ids": [
            component.id
            for account in _EBITDA_COMPONENTS
            for component in by_account[account]
        ],
    }
    return candidate, value, derivation, []


def get_qoe_bridge(
    session: Session,
    workspace_id: str,
    *,
    period_end: date | None = None,
    source_snapshot_id: str | None = None,
) -> dict:
    get_workspace_or_404(session, workspace_id)
    target = _target_for_workspace(session, workspace_id)
    if source_snapshot_id:
        snapshot = session.get(SourceSnapshot, source_snapshot_id)
        if snapshot is None or snapshot.workspace_id != workspace_id:
            raise UnderwritingDataError("source_snapshot_id does not belong to this workspace")

    base_fact, reported, derivation, warnings = _find_reported_ebitda(
        session, workspace_id, period_end, source_snapshot_id
    )
    bridge_period = base_fact.period_end if base_fact else period_end
    bridge_currency = base_fact.currency if base_fact else None
    zero = Decimal("0")
    if reported is None or bridge_period is None:
        return {
            "workspace_id": workspace_id,
            "target_id": target.id,
            "period_end": bridge_period,
            "currency": bridge_currency,
            "status": "incomplete",
            "reported_ebitda": None,
            "management_adjustments": zero,
            "management_ebitda": None,
            "sponsor_adjustments": zero,
            "sponsor_ebitda": None,
            "covenant_adjustments": zero,
            "covenant_ebitda": None,
            "included_adjustment_ids": [],
            "excluded_adjustment_count": 0,
            "source_snapshot_id": base_fact.source_snapshot_id if base_fact else source_snapshot_id,
            "source_locator": base_fact.source_locator if base_fact else None,
            "derivation": derivation,
            "warnings": warnings,
        }

    adjustments = list(
        session.scalars(
            select(QoEAdjustment).where(
                QoEAdjustment.workspace_id == workspace_id,
                QoEAdjustment.period_end == bridge_period,
            )
        )
    )
    approved = [
        adjustment
        for adjustment in adjustments
        if adjustment.status == "approved" and adjustment.currency == bridge_currency
    ]
    currency_excluded = [
        adjustment
        for adjustment in adjustments
        if adjustment.status == "approved" and adjustment.currency != bridge_currency
    ]
    if currency_excluded:
        warnings.append("Approved adjustments in a different currency were excluded")

    totals = {
        layer: sum(
            (adjustment.amount for adjustment in approved if adjustment.bridge_layer == layer),
            zero,
        )
        for layer in ("management", "sponsor", "covenant")
    }
    management = reported + totals["management"]
    sponsor = management + totals["sponsor"]
    covenant = sponsor + totals["covenant"]
    return {
        "workspace_id": workspace_id,
        "target_id": target.id,
        "period_end": bridge_period,
        "currency": bridge_currency,
        "status": "ready",
        "reported_ebitda": reported,
        "management_adjustments": totals["management"],
        "management_ebitda": management,
        "sponsor_adjustments": totals["sponsor"],
        "sponsor_ebitda": sponsor,
        "covenant_adjustments": totals["covenant"],
        "covenant_ebitda": covenant,
        "included_adjustment_ids": [adjustment.id for adjustment in approved],
        "excluded_adjustment_count": len(adjustments) - len(approved),
        "source_snapshot_id": base_fact.source_snapshot_id,
        "source_locator": base_fact.source_locator,
        "derivation": derivation,
        "warnings": warnings,
    }


def _validate_source_ids(session: Session, workspace_id: str, source_ids: list[str]) -> None:
    if len(source_ids) != len(set(source_ids)):
        raise UnderwritingDataError("source_snapshot_ids cannot contain duplicates")
    if not source_ids:
        return
    found = set(
        session.scalars(
            select(SourceSnapshot.id).where(
                SourceSnapshot.workspace_id == workspace_id,
                SourceSnapshot.id.in_(source_ids),
            )
        )
    )
    missing = set(source_ids) - found
    if missing:
        raise UnderwritingDataError(
            "Unknown source snapshots for this workspace: " + ", ".join(sorted(missing))
        )


def create_analysis_run(
    session: Session, workspace_id: str, data: AnalysisRunCreate
) -> AnalysisRun:
    get_workspace_or_404(session, workspace_id)
    _validate_source_ids(session, workspace_id, data.source_snapshot_ids)
    completed_at = data.completed_at or now_utc()
    started_at = data.started_at or completed_at
    version, latest = _next_stream_version(
        session, AnalysisRun, workspace_id, AnalysisRun.run_type, data.run_type
    )
    input_manifest = _json_safe(data.input_manifest)
    output_summary = _json_safe(data.output_summary)
    run_input_hash = content_hash(
        {
            "input_manifest": input_manifest,
            "source_snapshot_ids": sorted(data.source_snapshot_ids),
            "model_version": data.model_version,
            "prompt_version": data.prompt_version,
            "code_version": data.code_version,
        }
    )
    run_content_hash = content_hash(
        {
            "status": data.status,
            "output_summary": output_summary,
            "error_message": data.error_message,
        }
    )
    if data.input_hash is not None and data.input_hash != run_input_hash:
        raise UnderwritingDataError("input_hash does not match the server-canonical run inputs")
    if data.content_hash is not None and data.content_hash != run_content_hash:
        raise UnderwritingDataError("content_hash does not match the server-canonical run output")
    run = AnalysisRun(
        workspace_id=workspace_id,
        run_type=data.run_type,
        version=version,
        supersedes_id=latest.id if latest else None,
        status=data.status,
        input_hash=run_input_hash,
        content_hash=run_content_hash,
        source_snapshot_ids=data.source_snapshot_ids,
        input_manifest=input_manifest,
        output_summary=output_summary,
        model_version=data.model_version,
        prompt_version=data.prompt_version,
        code_version=data.code_version,
        error_message=data.error_message,
        created_by=data.created_by,
        started_at=started_at,
        completed_at=completed_at,
    )
    session.add(run)
    session.commit()
    session.refresh(run)
    return run


def list_analysis_runs(session: Session, workspace_id: str) -> list[AnalysisRun]:
    get_workspace_or_404(session, workspace_id)
    return list(
        session.scalars(
            select(AnalysisRun)
            .where(AnalysisRun.workspace_id == workspace_id)
            .order_by(AnalysisRun.created_at.desc())
        )
    )


def create_artifact_version(
    session: Session, workspace_id: str, data: ArtifactVersionCreate
) -> ArtifactVersion:
    get_workspace_or_404(session, workspace_id)
    _validate_source_ids(session, workspace_id, data.source_snapshot_ids)
    if data.analysis_run_id:
        run = session.get(AnalysisRun, data.analysis_run_id)
        if run is None or run.workspace_id != workspace_id:
            raise UnderwritingDataError("analysis_run_id does not belong to this workspace")

    version, latest = _next_stream_version(
        session, ArtifactVersion, workspace_id, ArtifactVersion.artifact_type, data.artifact_type
    )
    content_payload = (
        data.content_json
        if data.content_json is not None
        else data.content_text
        if data.content_text is not None
        else {"file_uri": data.file_uri}
    )
    if data.file_uri is not None and data.content_json is None and data.content_text is None:
        raise UnderwritingDataError(
            "file_uri-only artifacts cannot be content-verified; provide server-readable content"
        )
    artifact_input_hash = content_hash(
        {
            "input_manifest": _json_safe(data.input_manifest),
            "analysis_run_id": data.analysis_run_id,
            "source_snapshot_ids": sorted(data.source_snapshot_ids),
        }
    )
    artifact_content_hash = content_hash(content_payload)
    if data.input_hash is not None and data.input_hash != artifact_input_hash:
        raise UnderwritingDataError(
            "input_hash does not match the server-canonical artifact inputs"
        )
    if data.content_hash is not None and data.content_hash != artifact_content_hash:
        raise UnderwritingDataError(
            "content_hash does not match the server-canonical artifact content"
        )
    artifact = ArtifactVersion(
        workspace_id=workspace_id,
        artifact_type=data.artifact_type,
        version=version,
        supersedes_id=latest.id if latest else None,
        analysis_run_id=data.analysis_run_id,
        source_snapshot_ids=data.source_snapshot_ids,
        input_hash=artifact_input_hash,
        content_hash=artifact_content_hash,
        content_json=_json_safe(data.content_json),
        content_text=data.content_text,
        file_uri=data.file_uri,
        artifact_metadata=_json_safe(data.artifact_metadata),
        created_by=data.created_by,
    )
    session.add(artifact)
    session.commit()
    session.refresh(artifact)
    return artifact


def list_artifact_versions(session: Session, workspace_id: str) -> list[ArtifactVersion]:
    get_workspace_or_404(session, workspace_id)
    return list(
        session.scalars(
            select(ArtifactVersion)
            .where(ArtifactVersion.workspace_id == workspace_id)
            .order_by(ArtifactVersion.created_at.desc())
        )
    )


__all__ = [
    "UnderwritingDataConflict",
    "UnderwritingDataError",
    "content_hash",
    "create_account_mapping",
    "create_analysis_run",
    "create_artifact_version",
    "create_private_target",
    "create_qoe_adjustment",
    "decide_qoe_adjustment",
    "get_qoe_bridge",
    "import_financial_rows",
    "list_account_mappings",
    "list_analysis_runs",
    "list_artifact_versions",
    "list_financial_facts",
    "list_import_exceptions",
    "list_qoe_adjustments",
    "list_reconciliations",
    "list_source_snapshots",
    "normalize_account",
    "parse_financial_csv",
    "parse_financial_xlsx",
    "register_source_snapshot",
    "resolve_import_exception",
]

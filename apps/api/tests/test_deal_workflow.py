"""Focused offline coverage for the institutional deal workflow and IC domain."""
from __future__ import annotations

import hashlib
import io
import json

import pytest
from fastapi import FastAPI, Request
from fastapi.testclient import TestClient
from docx import Document
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from src.db.base import Base
from src.db.session import get_session
from src.models.deal_workflow import (
    DealStageTransition,
    ICPacketExport,
    Organization,
    WorkflowAuditEvent,
)
from src.models.underwriting_model import UnderwritingCaseDecision, UnderwritingCaseVersion
from src.models.workspace import Workspace
from src.routers.deal_workflow import router
from src.schemas.deal_workflow import (
    ActorContext,
    ConditionPatch,
    DealCreate,
    DealPatch,
    DiligenceAttachmentCreate,
    DiligenceRequestCreate,
    DiligenceResponseCreate,
    DiligenceReview,
    ExportRequest,
    FundCreate,
    ICCommentCreate,
    ICCommentResolve,
    ICDecisionCreate,
    ICPacketCreate,
    LedgerEntryCreate,
    LedgerEntryRevision,
    MilestoneCreate,
    OrganizationCreate,
    StageGateResolve,
    StageTransitionCreate,
    TaskCreate,
    TaskPatch,
    TeamMemberCreate,
    WorkstreamCreate,
)
from src.schemas.identity import PrincipalContext
from src.services import deal_workflow_service as service
from src.services import evidence_service
from src.services import ic_export_service


@pytest.fixture()
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as session:
        yield session
    engine.dispose()


def _setup(db: Session, *, suffix: str = "one"):
    creator = ActorContext(actor_id=f"lead-{suffix}", display_name="Deal Lead")
    organization = service.create_organization(
        db, OrganizationCreate(name=f"Test Organization {suffix}", slug=f"test-org-{suffix}"), creator
    )
    actor = creator.model_copy(update={"organization_id": organization.id})
    fund = service.create_fund(db, organization.id, FundCreate(name="Fund I"), actor)
    deal = service.create_deal(
        db,
        fund.id,
        DealCreate(
            code=f"P-{suffix}",
            name=f"Project {suffix.title()}",
            target_company=f"Target {suffix.title()}",
        ),
        actor,
    )
    return actor, organization, fund, deal


def _advance_to_ic(db: Session, deal_id: str, actor: ActorContext) -> None:
    stages = ["sourcing", "screening", "initial_review", "diligence"]
    next_stage = {
        "sourcing": "screening",
        "screening": "initial_review",
        "initial_review": "diligence",
        "diligence": "ic_review",
    }
    for stage in stages:
        for gate in service.list_gates(db, deal_id, actor, stage):
            if gate.status == "pending":
                service.resolve_gate(
                    db,
                    gate.id,
                    StageGateResolve(status="satisfied", resolution_note="Reviewed"),
                    actor,
                )
        service.transition_deal(
            db,
            deal_id,
            StageTransitionCreate(to_stage=next_stage[stage], rationale="Stage review complete"),
            actor,
        )
    for gate in service.list_gates(db, deal_id, actor, "ic_review"):
        service.resolve_gate(
            db,
            gate.id,
            StageGateResolve(status="satisfied", resolution_note="Materials checked"),
            actor,
        )


def _packet(
    db: Session,
    deal,
    previous_packet_id: str | None = None,
    *,
    revenue: int = 100,
) -> ICPacketCreate:
    if deal.workspace_id is None:
        workspace = Workspace(
            name=f"{deal.name} governed workspace",
            organization_id=deal.organization_id,
            deal_type="buyout",
            status="draft",
        )
        db.add(workspace)
        db.flush()
        deal.workspace_id = workspace.id
        db.commit()
    prior_cases = list(
        db.scalars(
            select(UnderwritingCaseVersion).where(
                UnderwritingCaseVersion.workspace_id == deal.workspace_id,
                UnderwritingCaseVersion.case_key == "base",
            )
        )
    )
    version = len(prior_cases) + 1
    digest = hashlib.sha256(f"{deal.id}:{version}:{revenue}".encode()).hexdigest()
    case = UnderwritingCaseVersion(
        workspace_id=deal.workspace_id,
        case_key="base",
        label="Base case",
        version=version,
        schema_version="1.0",
        assumptions={"revenue": revenue, "entry_enterprise_value": 250_000_000},
        result={"irr": 0.24, "moic": 2.5},
        input_hash=digest,
        output_hash=hashlib.sha256(f"result:{digest}".encode()).hexdigest(),
        created_by="model-analyst",
        change_note="Governed workflow fixture",
    )
    db.add(case)
    db.commit()
    db.add(
        UnderwritingCaseDecision(
            workspace_id=deal.workspace_id,
            case_version_id=case.id,
            decision="submitted",
            actor="model-analyst",
            rationale="Submitted for review",
        )
    )
    db.commit()
    db.add(
        UnderwritingCaseDecision(
            workspace_id=deal.workspace_id,
            case_version_id=case.id,
            decision="approved",
            actor="investment-partner",
            rationale="Model tie-out complete",
        )
    )
    db.commit()
    evidence = evidence_service.create(
        db,
        deal.workspace_id,
        claim=f"Governed base-case revenue is {revenue}.",
        claim_type="assumption",
        source_name="Underwriting model fixture",
        source_type="analyst_model",
        evidence_text=f"Base-case revenue assumption: {revenue}.",
        confidence=0.9,
        agent_name="model-analyst",
    )
    db.commit()
    service.create_ledger_entry(
        db,
        deal.id,
        LedgerEntryCreate(
            entry_type="thesis",
            title=f"Governed base case v{version}",
            description="Approved underwriting case supports the investment thesis.",
            status="validated",
            evidence_refs=[evidence.ref],
        ),
        ActorContext(actor_id="model-analyst", organization_id=deal.organization_id),
    )
    return ICPacketCreate(
        title="Investment Committee Memorandum",
        case_version_ids=[case.id],
        workspace_evidence_refs=[evidence.ref],
        decision_request={"ask": "Approve signing authority", "equity_commitment": 125_000_000},
        previous_packet_id=previous_packet_id,
    )


def test_tenant_scope_and_optimistic_deal_update(db: Session):
    actor, organization, _, deal = _setup(db)
    outsider = ActorContext(actor_id="outside", organization_id="f" * 32)

    # Cross-tenant reads get the SAME 404 as an unknown id — a 403 would confirm the deal id
    # exists in another tenant (existence oracle).
    with pytest.raises(service.NotFound):
        service.get_deal(db, deal.id, outsider)

    assert [item.id for item in service.list_organizations(db, actor)] == [organization.id]
    deal.workspace_id = "a" * 32
    db.commit()
    assert service.get_deal_by_workspace(db, "a" * 32, actor).id == deal.id

    updated = service.update_deal(
        db,
        deal.id,
        DealPatch(expected_version=1, summary="Platform carve-out with recurring revenue"),
        actor,
    )
    assert updated.version == 2
    assert updated.organization_id == organization.id

    with pytest.raises(service.WorkflowConflict, match="version changed"):
        service.update_deal(
            db, deal.id, DealPatch(expected_version=1, summary="stale write"), actor
        )


def test_deal_workspace_links_bind_unowned_and_reject_cross_tenant(db: Session):
    actor, organization, fund, deal = _setup(db, suffix="workspace-link")
    other = service.create_organization(
        db,
        OrganizationCreate(name="Other Workspace Org", slug="other-workspace-org"),
        ActorContext(actor_id="other-lead"),
    )
    create_workspace = Workspace(name="Create link", deal_type="buyout", status="draft")
    update_workspace = Workspace(name="Update link", deal_type="buyout", status="draft")
    foreign_workspace = Workspace(
        name="Foreign link",
        organization_id=other.id,
        deal_type="buyout",
        status="draft",
    )
    db.add_all([create_workspace, update_workspace, foreign_workspace])
    db.commit()

    linked = service.create_deal(
        db,
        fund.id,
        DealCreate(
            code="LINK-CREATE",
            name="Create-time workspace link",
            target_company="Linked Target",
            workspace_id=create_workspace.id,
        ),
        actor,
    )
    assert linked.workspace_id == create_workspace.id
    assert db.get(Workspace, create_workspace.id).organization_id == organization.id

    updated = service.update_deal(
        db,
        deal.id,
        DealPatch(expected_version=1, workspace_id=update_workspace.id),
        actor,
    )
    assert updated.workspace_id == update_workspace.id
    assert db.get(Workspace, update_workspace.id).organization_id == organization.id

    with pytest.raises(service.WorkflowForbidden, match="different organization"):
        service.create_deal(
            db,
            fund.id,
            DealCreate(
                code="LINK-FOREIGN",
                name="Cross-tenant create",
                target_company="Foreign Target",
                workspace_id=foreign_workspace.id,
            ),
            actor,
        )
    with pytest.raises(service.WorkflowForbidden, match="different organization"):
        service.update_deal(
            db,
            deal.id,
            DealPatch(expected_version=2, workspace_id=foreign_workspace.id),
            actor,
        )
    assert db.get(Workspace, foreign_workspace.id).organization_id == other.id
    assert db.get(type(deal), deal.id).version == 2


def test_stage_gates_enforce_order_and_transition_log_is_append_only(db: Session):
    actor, _, _, deal = _setup(db, suffix="gates")

    with pytest.raises(service.WorkflowConflict, match="stage gates"):
        service.transition_deal(
            db, deal.id, StageTransitionCreate(to_stage="screening"), actor
        )
    with pytest.raises(service.WorkflowConflict, match="next stage"):
        service.transition_deal(
            db, deal.id, StageTransitionCreate(to_stage="diligence"), actor
        )

    sourcing_gate = service.list_gates(db, deal.id, actor, "sourcing")[0]
    service.resolve_gate(
        db,
        sourcing_gate.id,
        StageGateResolve(
            status="waived", resolution_note="Mandate exception approved", evidence_refs=["EV-9"]
        ),
        actor,
    )
    transition = service.transition_deal(
        db,
        deal.id,
        StageTransitionCreate(to_stage="screening", rationale="Screening authorized"),
        actor,
    )
    assert transition.sequence == 1
    assert service.get_deal(db, deal.id, actor).stage == "screening"

    transition.rationale = "tamper"
    with pytest.raises(ValueError, match="append-only"):
        db.commit()
    db.rollback()
    persisted = db.get(DealStageTransition, transition.id)
    assert persisted.rationale == "Screening authorized"


def test_execution_objects_enforce_task_dependencies(db: Session):
    actor, _, _, deal = _setup(db, suffix="tasks")
    workstream = service.create_workstream(
        db,
        deal.id,
        WorkstreamCreate(slug="financial", label="Financial diligence", lead_actor_id=actor.actor_id),
        actor,
    )
    milestone = service.create_milestone(
        db,
        deal.id,
        MilestoneCreate(workstream_id=workstream.id, title="QoE sign-off"),
        actor,
    )
    source_task = service.create_task(
        db,
        deal.id,
        TaskCreate(
            workstream_id=workstream.id,
            milestone_id=milestone.id,
            title="Reconcile historical EBITDA",
            priority="high",
        ),
        actor,
    )
    dependent = service.create_task(
        db,
        deal.id,
        TaskCreate(
            workstream_id=workstream.id,
            milestone_id=milestone.id,
            title="Approve sponsor EBITDA",
            dependency_task_ids=[source_task.id],
        ),
        actor,
    )

    with pytest.raises(service.WorkflowConflict, match="dependencies"):
        service.update_task(db, dependent.id, TaskPatch(status="complete"), actor)
    service.update_task(db, source_task.id, TaskPatch(status="complete"), actor)
    completed = service.update_task(db, dependent.id, TaskPatch(status="complete"), actor)
    assert completed.completed_by_actor_id == actor.actor_id
    assert len(service.list_tasks(db, deal.id, actor, status="complete")) == 2


def test_diligence_request_requires_independent_acceptance_and_tracks_artifact(db: Session):
    actor, organization, _, deal = _setup(db, suffix="requests")
    request = service.create_diligence_request(
        db,
        deal.id,
        DiligenceRequestCreate(
            title="Customer cohort detail",
            question="Provide monthly ARR by customer and cohort.",
            priority="high",
            send_now=True,
        ),
        actor,
    )
    respondent = ActorContext(actor_id="management-cfo", organization_id=organization.id)
    response = service.add_diligence_response(
        db,
        request.id,
        DiligenceResponseCreate(response_text="Uploaded the requested ARR cohort file."),
        respondent,
    )
    digest = hashlib.sha256(b"immutable workbook bytes").hexdigest()
    attachment = service.add_diligence_attachment(
        db,
        request.id,
        DiligenceAttachmentCreate(
            response_id=response.id,
            filename="ARR cohorts.xlsx",
            object_key=f"quarantine/{deal.id}/arr-cohorts-v1",
            source_hash=digest,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            size_bytes=26_144,
        ),
        respondent,
    )
    assert attachment.source_hash == digest

    with pytest.raises(service.WorkflowConflict, match="cannot accept"):
        service.review_diligence_request(
            db, request.id, DiligenceReview(action="accept", note="Self review"), respondent
        )
    accepted = service.review_diligence_request(
        db, request.id, DiligenceReview(action="accept", note="Tie-out completed"), actor
    )
    assert accepted.status == "accepted"
    assert accepted.accepted_by_actor_id == actor.actor_id


def test_ledger_revisions_preserve_history(db: Session):
    actor, _, _, deal = _setup(db, suffix="ledger")
    original = service.create_ledger_entry(
        db,
        deal.id,
        LedgerEntryCreate(
            entry_type="thesis",
            title="Mission-critical workflow",
            description="Low churn supports a resilient base case.",
            evidence_refs=["EV-001"],
        ),
        actor,
    )
    revision = service.revise_ledger_entry(
        db,
        original.id,
        LedgerEntryRevision(
            description="Validated cohort data supports a resilient base case.",
            status="validated",
            evidence_refs=["EV-001", "EV-004"],
        ),
        actor,
    )
    db.refresh(original)
    assert original.status == "superseded"
    assert revision.version == 2
    assert revision.root_entry_id == original.id
    assert revision.supersedes_entry_id == original.id
    assert service.list_ledger_entries(db, deal.id, actor) == [revision]
    assert len(service.list_ledger_entries(db, deal.id, actor, include_superseded=True)) == 2


def test_ic_readiness_freeze_four_eyes_conditions_and_export_manifest(db: Session):
    actor, organization, _, deal = _setup(db, suffix="ic")
    partner = ActorContext(
        actor_id="partner-1", display_name="Investment Partner", organization_id=organization.id
    )
    reviewer = ActorContext(
        actor_id="reviewer-1", display_name="IC Reviewer", organization_id=organization.id
    )
    service.add_team_member(
        db,
        deal.id,
        TeamMemberCreate(
            actor_id=partner.actor_id or "partner-1",
            display_name="Investment Partner",
            role="investment_partner",
        ),
        actor,
    )
    _advance_to_ic(db, deal.id, actor)
    packet = service.create_ic_packet(db, deal.id, _packet(db, deal), actor)

    readiness = service.evaluate_ic_readiness(db, packet.id, actor)
    assert readiness["ready"] is True
    submitted = service.submit_ic_packet(db, packet.id, actor)
    assert submitted.status == "submitted"
    assert submitted.frozen_at is not None

    comment = service.add_ic_comment(
        db,
        packet.id,
        ICCommentCreate(
            section_path="$.model_snapshot.irr",
            body="Tie the downside IRR to the debt schedule.",
            blocking=True,
        ),
        reviewer,
    )
    with pytest.raises(service.WorkflowConflict, match="second actor"):
        service.resolve_ic_comment(
            db, comment.id, ICCommentResolve(resolution="Resolved by author"), reviewer
        )
    service.resolve_ic_comment(
        db,
        comment.id,
        ICCommentResolve(resolution="Debt schedule tie-out attached as EV-015."),
        partner,
    )

    with pytest.raises(service.WorkflowConflict, match="submitter"):
        service.record_ic_decision(
            db,
            packet.id,
            ICDecisionCreate(decision="approve", rationale="Approved"),
            actor,
        )
    decision, conditions = service.record_ic_decision(
        db,
        packet.id,
        ICDecisionCreate(
            decision="conditional",
            rationale="Proceed subject to customer renewal confirmation.",
            conditions=[
                {
                    "description": "Confirm renewal of the top customer contract.",
                    "owner_actor_id": actor.actor_id,
                }
            ],
        ),
        partner,
    )
    assert decision.is_final is True
    assert len(conditions) == 1
    resolved = service.update_condition(
        db,
        conditions[0].id,
        ConditionPatch(
            status="satisfied",
            resolution_note="Executed renewal received.",
            evidence_refs=["EV-020"],
        ),
        actor,
    )
    assert resolved.status == "satisfied"

    export = service.create_export_manifest(
        db, packet.id, ExportRequest(format="pdf"), partner
    )
    assert export.manifest["packet"]["content_hash"] == packet.content_hash
    assert export.manifest["decisions"][0]["decision"] == "conditional"
    assert len(export.manifest_hash) == 64
    assert any(
        event.action == "ic_decision.recorded"
        for event in service.list_audit_events(db, deal.id, actor)
    )


def test_frozen_ic_packet_exports_real_json_xlsx_docx_and_pdf_files(db: Session):
    actor, _, _, deal = _setup(db, suffix="files")
    _advance_to_ic(db, deal.id, actor)
    packet = service.create_ic_packet(db, deal.id, _packet(db, deal), actor)
    service.submit_ic_packet(db, packet.id, actor)

    exported = {
        format_name: ic_export_service.render_and_record_export(
            db, packet.id, ExportRequest(format=format_name), actor
        )
        for format_name in ("json", "xlsx", "docx", "pdf")
    }
    payload = json.loads(exported["json"].content)
    assert payload["export_metadata"]["packet_id"] == packet.id
    assert payload["export_metadata"]["packet_content_hash"] == packet.content_hash

    workbook = load_workbook(io.BytesIO(exported["xlsx"].content), read_only=True, data_only=False)
    assert {"Export Metadata", "Model Snapshot", "Evidence Manifest"}.issubset(workbook.sheetnames)
    document = Document(io.BytesIO(exported["docx"].content))
    assert any("Investment Committee Pack" in paragraph.text for paragraph in document.paragraphs)
    assert exported["pdf"].content.startswith(b"%PDF-")

    records = list(db.scalars(select(ICPacketExport).where(ICPacketExport.packet_id == packet.id)))
    assert len(records) == 4
    for record in records:
        result = exported[record.format]
        assert record.manifest["file_sha256"] == hashlib.sha256(result.content).hexdigest()
        assert record.manifest["packet_content_hash"] == packet.content_hash


def test_readiness_reports_blockers_and_packet_diff_is_path_specific(db: Session):
    actor, _, _, deal = _setup(db, suffix="diff")
    first = service.create_ic_packet(db, deal.id, _packet(db, deal, revenue=100), actor)
    readiness = service.evaluate_ic_readiness(db, first.id, actor)
    failed = {item["code"] for item in readiness["checks"] if not item["passed"]}
    assert {"deal_stage", "required_gates"}.issubset(failed)

    second = service.create_ic_packet(
        db, deal.id, _packet(db, deal, previous_packet_id=first.id, revenue=115), actor
    )
    result = service.diff_ic_packets(db, first.id, second.id, actor)
    changes = {item["path"]: item for item in result["changes"]}
    assert changes["$.scenario_snapshot.cases[0].assumptions.revenue"] == {
        "path": "$.scenario_snapshot.cases[0].assumptions.revenue",
        "change": "changed",
        "before": 100,
        "after": 115,
    }


def test_audit_events_are_append_only(db: Session):
    actor, _, _, deal = _setup(db, suffix="audit")
    event = db.scalar(
        select(WorkflowAuditEvent).where(WorkflowAuditEvent.deal_id == deal.id)
    )
    assert event is not None
    event.action = "tampered"
    with pytest.raises(ValueError, match="append-only"):
        db.commit()
    db.rollback()
    assert db.get(WorkflowAuditEvent, event.id).action == "deal.created"


def test_router_translates_tenant_and_version_conflicts():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    app = FastAPI()
    app.include_router(router)

    def session_override():
        with Session(engine, expire_on_commit=False) as session:
            yield session

    app.dependency_overrides[get_session] = session_override
    with TestClient(app) as client:
        org_response = client.post(
            "/api/organizations", json={"name": "API Organization", "slug": "api-org"}
        )
        assert org_response.status_code == 201, org_response.text
        organization_id = org_response.json()["id"]
        headers = {
            "X-Actor-ID": "api-lead",
            "X-Actor-Name": "API Lead",
            "X-Organization-ID": organization_id,
        }
        fund_response = client.post(
            f"/api/organizations/{organization_id}/funds",
            json={"name": "API Fund"},
            headers=headers,
        )
        assert fund_response.status_code == 201, fund_response.text
        deal_response = client.post(
            f"/api/funds/{fund_response.json()['id']}/deals",
            json={"code": "API-1", "name": "API Deal", "target_company": "API Target"},
            headers=headers,
        )
        assert deal_response.status_code == 201, deal_response.text
        deal_id = deal_response.json()["id"]

        # Cross-tenant reads answer with the SAME 404 as an unknown id — a 403 would confirm
        # the deal exists in another tenant (existence oracle).
        cross_tenant = client.get(
            f"/api/deals/{deal_id}",
            headers={"X-Actor-ID": "outside", "X-Organization-ID": "f" * 32},
        )
        assert cross_tenant.status_code == 404
        missing = client.get(
            "/api/deals/" + "0" * 32,
            headers={"X-Actor-ID": "outside", "X-Organization-ID": "f" * 32},
        )
        assert missing.status_code == 404

        first_update = client.patch(
            f"/api/deals/{deal_id}",
            json={"expected_version": 1, "summary": "First write"},
            headers=headers,
        )
        assert first_update.status_code == 200
        stale_update = client.patch(
            f"/api/deals/{deal_id}",
            json={"expected_version": 1, "summary": "Stale write"},
            headers=headers,
        )
        assert stale_update.status_code == 409
    engine.dispose()


def test_export_verification_endpoint_reports_valid_and_rehashed_tampering():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as db:
        actor, organization, _, deal = _setup(db, suffix="verify-api")
        _advance_to_ic(db, deal.id, actor)
        packet = service.create_ic_packet(db, deal.id, _packet(db, deal), actor)
        service.submit_ic_packet(db, packet.id, actor)
        export = service.create_export_manifest(
            db, packet.id, ExportRequest(format="json"), actor
        )
        export_id = export.id
        headers = {
            "X-Actor-ID": actor.actor_id or "verify-lead",
            "X-Organization-ID": organization.id,
        }

    app = FastAPI()
    app.include_router(router)

    def session_override():
        with Session(engine, expire_on_commit=False) as session:
            yield session

    app.dependency_overrides[get_session] = session_override
    with TestClient(app) as client:
        valid = client.get(f"/api/ic-exports/{export_id}/verification", headers=headers)
        assert valid.status_code == 200, valid.text
        assert valid.json()["valid"] is True

        with Session(engine, expire_on_commit=False) as db:
            stored = db.get(ICPacketExport, export_id)
            assert stored is not None
            tampered = json.loads(json.dumps(stored.manifest))
            tampered["packet"]["content_hash"] = "f" * 64
            stored.manifest = tampered
            stored.manifest_hash = service._sha256(tampered)
            db.commit()

        invalid = client.get(f"/api/ic-exports/{export_id}/verification", headers=headers)
        assert invalid.status_code == 200, invalid.text
        result = invalid.json()
        assert result["valid"] is False
        checks = {item["code"]: item["passed"] for item in result["checks"]}
        assert checks["canonical_manifest_hash"] is True
        assert checks["packet_content_binding"] is False
    engine.dispose()


def test_authenticated_principal_cannot_use_legacy_organization_provisioning():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    app = FastAPI()

    @app.middleware("http")
    async def verified_principal(request: Request, call_next):
        request.state.principal = PrincipalContext(
            user_id="verified-user",
            session_id="verified-session",
            email="verified@example.com",
            display_name="Verified User",
            organization_id="a" * 32,
            membership_id="verified-membership",
            role="owner",
        )
        return await call_next(request)

    app.include_router(router)

    def session_override():
        with Session(engine, expire_on_commit=False) as session:
            yield session

    app.dependency_overrides[get_session] = session_override
    with TestClient(app) as client:
        response = client.post(
            "/api/organizations",
            json={"name": "Orphan Organization", "slug": "orphan-org"},
        )
        assert response.status_code == 403
        assert "registration flow" in response.json()["detail"]
    with Session(engine) as db:
        assert db.scalar(select(Organization)) is None
    engine.dispose()
